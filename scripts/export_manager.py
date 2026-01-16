"""
Export Manager
==============
Export job data and reports to various formats.

Usage:
    python export_manager.py --excel              # Export to Excel
    python export_manager.py --pdf                # Generate PDF report
    python export_manager.py --csv                # Export to CSV
    python export_manager.py --all                # Export all formats
"""

import sys

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import argparse

# Project paths
PROJECT_ROOT = Path(__file__).parent.parent
DATA_DIR = PROJECT_ROOT / 'data' / 'raw'
REPORTS_DIR = PROJECT_ROOT / 'outputs' / 'reports'
EXPORTS_DIR = PROJECT_ROOT / 'data' / 'exports'

# Ensure directories exist
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


class ExportManager:
    """Export job data to various formats."""
    
    def __init__(self):
        self.jobs = []
        self.analysis = {}
        self._load_data()
    
    def _load_data(self):
        """Load job data and analysis."""
        # Load jobs
        try:
            import pandas as pd
            csv_files = list(DATA_DIR.glob('jobs_*.csv'))
            if csv_files:
                latest = max(csv_files, key=lambda f: f.stat().st_mtime)
                df = pd.read_csv(latest)
                self.jobs = df.to_dict('records')
        except ImportError:
            json_files = list(DATA_DIR.glob('jobs_*.json'))
            if json_files:
                latest = max(json_files, key=lambda f: f.stat().st_mtime)
                with open(latest, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.jobs = data if isinstance(data, list) else data.get('jobs', [])
        
        # Load analysis
        analysis_files = list(REPORTS_DIR.glob('analysis_*.json'))
        if analysis_files:
            latest = max(analysis_files, key=lambda f: f.stat().st_mtime)
            with open(latest, 'r', encoding='utf-8') as f:
                self.analysis = json.load(f)
        
        print(f"[i] Loaded {len(self.jobs)} jobs")
    
    def export_to_csv(self, output_path: str = None) -> str:
        """Export jobs to CSV file."""
        if not self.jobs:
            print("[X] No job data to export")
            return None
        
        try:
            import pandas as pd
        except ImportError:
            print("[X] pandas required: pip install pandas")
            return None
        
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = EXPORTS_DIR / f"jobs_export_{timestamp}.csv"
        
        df = pd.DataFrame(self.jobs)
        df.to_csv(output_path, index=False, encoding='utf-8')
        
        print(f"[OK] Exported {len(self.jobs)} jobs to: {output_path}")
        return str(output_path)
    
    def export_to_excel(self, output_path: str = None) -> str:
        """Export jobs to Excel with multiple sheets and formatting."""
        if not self.jobs:
            print("[X] No job data to export")
            return None
        
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            print("[X] Required: pip install pandas openpyxl")
            return None
        
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = EXPORTS_DIR / f"jobs_report_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0077B5", end_color="0077B5", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: All Jobs
        ws1 = wb.active
        ws1.title = "All Jobs"
        
        df = pd.DataFrame(self.jobs)
        # Select key columns
        columns = ['title', 'company', 'location', 'job_type', 'salary', 'url']
        available_cols = [c for c in columns if c in df.columns]
        if available_cols:
            df_export = df[available_cols]
        else:
            df_export = df
        
        for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        for col in ws1.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        # Sheet 2: Summary Statistics
        ws2 = wb.create_sheet("Summary")
        
        summary_data = [
            ["LinkedIn Job Analysis Report", ""],
            ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M')],
            ["", ""],
            ["Total Jobs", len(self.jobs)],
            ["Unique Companies", len(set(j.get('company', '') for j in self.jobs))],
            ["Unique Locations", len(set(j.get('location', '') for j in self.jobs))],
        ]
        
        # Add top skills if available
        if self.analysis.get('skills', {}).get('top_10'):
            summary_data.append(["", ""])
            summary_data.append(["Top Skills", "Count"])
            for skill, count in list(self.analysis['skills']['top_10'].items())[:10]:
                summary_data.append([skill.upper(), count])
        
        for r_idx, row in enumerate(summary_data, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif r_idx in [4, 5, 6, 8]:
                    cell.font = Font(bold=True)
        
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 20
        
        # Sheet 3: Companies
        ws3 = wb.create_sheet("Top Companies")
        
        company_counts = {}
        for job in self.jobs:
            company = job.get('company', 'Unknown')
            company_counts[company] = company_counts.get(company, 0) + 1
        
        sorted_companies = sorted(company_counts.items(), key=lambda x: x[1], reverse=True)[:20]
        
        ws3.cell(row=1, column=1, value="Company").font = header_font
        ws3.cell(row=1, column=1).fill = header_fill
        ws3.cell(row=1, column=2, value="Open Positions").font = header_font
        ws3.cell(row=1, column=2).fill = header_fill
        
        for r_idx, (company, count) in enumerate(sorted_companies, 2):
            ws3.cell(row=r_idx, column=1, value=company)
            ws3.cell(row=r_idx, column=2, value=count)
        
        ws3.column_dimensions['A'].width = 40
        ws3.column_dimensions['B'].width = 15
        
        # Sheet 4: Locations
        ws4 = wb.create_sheet("Locations")
        
        location_counts = {}
        for job in self.jobs:
            location = job.get('location', 'Unknown')
            location_counts[location] = location_counts.get(location, 0) + 1
        
        sorted_locations = sorted(location_counts.items(), key=lambda x: x[1], reverse=True)[:20]
        
        ws4.cell(row=1, column=1, value="Location").font = header_font
        ws4.cell(row=1, column=1).fill = header_fill
        ws4.cell(row=1, column=2, value="Jobs").font = header_font
        ws4.cell(row=1, column=2).fill = header_fill
        
        for r_idx, (location, count) in enumerate(sorted_locations, 2):
            ws4.cell(row=r_idx, column=1, value=location)
            ws4.cell(row=r_idx, column=2, value=count)
        
        ws4.column_dimensions['A'].width = 40
        ws4.column_dimensions['B'].width = 10
        
        # Save workbook
        wb.save(output_path)
        print(f"[OK] Excel report saved to: {output_path}")
        return str(output_path)
    
    def export_to_pdf(self, output_path: str = None) -> str:
        """Generate PDF report."""
        if not self.jobs:
            print("[X] No job data to export")
            return None
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            print("[X] Required: pip install reportlab")
            return None
        
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = EXPORTS_DIR / f"jobs_report_{timestamp}.pdf"
        
        doc = SimpleDocTemplate(str(output_path), pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#0077B5')
        )
        elements.append(Paragraph("LinkedIn Job Analysis Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Summary
        elements.append(Paragraph("Summary", styles['Heading2']))
        summary_data = [
            ["Metric", "Value"],
            ["Total Jobs", str(len(self.jobs))],
            ["Unique Companies", str(len(set(j.get('company', '') for j in self.jobs)))],
            ["Unique Locations", str(len(set(j.get('location', '') for j in self.jobs)))],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0077B5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Top Skills
        if self.analysis.get('skills', {}).get('top_10'):
            elements.append(Paragraph("Top In-Demand Skills", styles['Heading2']))
            
            skills_data = [["Skill", "Mentions"]]
            for skill, count in list(self.analysis['skills']['top_10'].items())[:10]:
                skills_data.append([skill.upper(), str(count)])
            
            skills_table = Table(skills_data, colWidths=[3*inch, 2*inch])
            skills_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0077B5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(skills_table)
            elements.append(Spacer(1, 20))
        
        # Top Companies
        elements.append(Paragraph("Top Hiring Companies", styles['Heading2']))
        
        company_counts = {}
        for job in self.jobs:
            company = job.get('company', 'Unknown')
            company_counts[company] = company_counts.get(company, 0) + 1
        
        sorted_companies = sorted(company_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        
        company_data = [["Company", "Open Positions"]]
        for company, count in sorted_companies:
            company_data.append([company[:40], str(count)])
        
        company_table = Table(company_data, colWidths=[4*inch, 1.5*inch])
        company_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0077B5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(company_table)
        
        # Build PDF
        doc.build(elements)
        print(f"[OK] PDF report saved to: {output_path}")
        return str(output_path)
    
    def export_to_json(self, output_path: str = None) -> str:
        """Export jobs to JSON file."""
        if not self.jobs:
            print("[X] No job data to export")
            return None
        
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = EXPORTS_DIR / f"jobs_export_{timestamp}.json"
        
        output = {
            'exported_at': datetime.now().isoformat(),
            'total_jobs': len(self.jobs),
            'jobs': self.jobs
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        
        print(f"[OK] Exported {len(self.jobs)} jobs to: {output_path}")
        return str(output_path)
    
    def export_all(self) -> Dict[str, str]:
        """Export to all formats."""
        results = {}
        
        print("\n" + "=" * 60)
        print("EXPORTING TO ALL FORMATS")
        print("=" * 60 + "\n")
        
        results['csv'] = self.export_to_csv()
        results['json'] = self.export_to_json()
        results['excel'] = self.export_to_excel()
        results['pdf'] = self.export_to_pdf()
        
        print("\n" + "=" * 60)
        print("EXPORT SUMMARY")
        print("=" * 60)
        for format_name, path in results.items():
            status = "[OK]" if path else "[FAIL]"
            print(f"  {status} {format_name.upper()}: {path or 'Failed'}")
        
        return results


def main():
    parser = argparse.ArgumentParser(description='Export Manager - Export job data to various formats')
    parser.add_argument('--csv', action='store_true', help='Export to CSV')
    parser.add_argument('--excel', action='store_true', help='Export to Excel')
    parser.add_argument('--pdf', action='store_true', help='Generate PDF report')
    parser.add_argument('--json', action='store_true', help='Export to JSON')
    parser.add_argument('--all', action='store_true', help='Export all formats')
    parser.add_argument('--output', '-o', type=str, help='Output file path')
    
    args = parser.parse_args()
    
    manager = ExportManager()
    
    if args.all:
        manager.export_all()
    elif args.csv:
        manager.export_to_csv(args.output)
    elif args.excel:
        manager.export_to_excel(args.output)
    elif args.pdf:
        manager.export_to_pdf(args.output)
    elif args.json:
        manager.export_to_json(args.output)
    else:
        parser.print_help()
        print("\nExamples:")
        print("  python export_manager.py --excel")
        print("  python export_manager.py --pdf")
        print("  python export_manager.py --all")


if __name__ == "__main__":
    main()
